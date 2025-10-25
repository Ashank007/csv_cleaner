import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

st.set_page_config(page_title="CSV Cleaner", layout="centered")

st.title("📘 CSV → Clean Excel Converter")

st.write("Upload your CSV file and get a formatted Excel file with the same name.")

# Columns to keep
COLUMNS_TO_KEEP = [
    "Name",
    "Email",
    "External Id",
    "Course",
    "Estimated Learning Hours",
    "Completed",
    "Course Grade"
]

uploaded_file = st.file_uploader("📂 Upload CSV file", type=["csv"])

def process_csv(file):
    df = pd.read_csv(file)

    # Filter only required columns (ignore missing)
    df_filtered = df[[col for col in COLUMNS_TO_KEEP if col in df.columns]]

    # Save to Excel (in-memory)
    output = BytesIO()
    df_filtered.to_excel(output, index=False)
    output.seek(0)

    # Open workbook for formatting
    wb = load_workbook(output)
    ws = wb.active

    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    # Header styling
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align

    # Center align all data
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center_align

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

    # Add Excel table
    last_row = ws.max_row
    last_col = ws.max_column
    end_col_letter = get_column_letter(last_col)
    table_ref = f"A1:{end_col_letter}{last_row}"

    table = Table(displayName="FilteredData", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save formatted Excel in-memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


if uploaded_file is not None:
    st.success("✅ File uploaded successfully!")

    # Extract base filename (without extension)
    base_filename = os.path.splitext(uploaded_file.name)[0]
    output_filename = f"{base_filename}.xlsx"

    if st.button("🧹 Clean & Convert"):
        try:
            processed_excel = process_csv(uploaded_file)
            st.success("🎉 File processed successfully!")

            st.download_button(
                label="⬇️ Download Cleaned Excel",
                data=processed_excel,
                file_name=output_filename,  # 👈 same as input file name
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"⚠️ Error: {e}")


